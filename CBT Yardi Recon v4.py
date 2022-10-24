import time
import openpyxl as xl
from itertools import groupby
from operator import itemgetter
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from openpyxl import Workbook
from datetime import date
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font, Side, Border, Alignment

start_time = time.time()
current_date = date.today()


def set_alignment(cell_range, position):  # Style position of cells.
    for rows in ws3[cell_range]:
        for cells in rows:
            cells.alignment = Alignment(horizontal=position)


def set_bold(cell_range):
    for rows in ws3[cell_range]:
        for cells in rows:
            cells.font = Font(bold=True, size=11)


def set_border(cell_range):
    thin = Side(border_style="thin", color="000000")
    for rows in ws3[cell_range]:
        for cells in rows:
            cells.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def red():
    red_font = Font(size=10, color='FF0000')
    ws3.conditional_formatting.add(f'A1:{get_column_letter(ws3.max_column)}{ws3.max_row}',
                                   CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=red_font))


yellowFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

month = str(current_date.month)
if current_date.month < 10:
    month = '0' + str(current_date.month)

year = str(current_date.year)
date = f'{month}/{str(current_date.day)}/{str(current_date.year)}'

# Run CBTYardiCleanUP first to filter out of CBT and Yardi deposit sheets. That program also sorts the CBT by 'Account Name' column alphabetically.
# Make sure to change path place for new users, ctrl+r, '\vpham\'.

CBT_Deposits_path = r'C:\Users\vpham\PycharmProjects\pythonProject\Programs\CBT Yardi Recon\Deposit excel files\CBT Deposits.xlsx'
Yardi_Deposits_path = r'C:\Users\vpham\PycharmProjects\pythonProject\Programs\CBT Yardi Recon\Deposit excel files\Yardi Deposits.xlsx'
CBT_Yardi_Reference_path = r'L:\Programs\CBT Yardi Recon\CBT Yardi Reference.xlsx'
wb1 = xl.load_workbook(CBT_Deposits_path)
wb2 = xl.load_workbook(Yardi_Deposits_path)
wb4 = xl.load_workbook(CBT_Yardi_Reference_path)

# Make sure that all the Excel files have only 1 sheet.
ws1 = wb1.worksheets[0]
ws2 = wb2.worksheets[0]
ws4 = wb4.worksheets[0]

propertynamelist = []
cbtaccnumlist = []
yardicodelist = []

for i in range(2, ws4.max_row + 1):
    propertynamelist.append(ws4.cell(row=i, column=1).value)
    cbtaccnumlist.append(ws4.cell(row=i, column=2).value)
    yardicodelist.append(ws4.cell(row=i, column=3).value)

cbtaccnumliststr = [str(x) for x in cbtaccnumlist]

cbtaccnumlist_tuple = list(tuple(zip(cbtaccnumliststr, propertynamelist)))
yardicodelist_tuple = list(tuple(zip(yardicodelist, propertynamelist)))

CBT_value = []
CBT_key = []
ledger_date = []
ref_text = []


for i in range(2, ws1.max_row + 1):
    CBT_key.append(ws1.cell(row=i, column=1).value)
    CBT_value.append(ws1.cell(row=i, column=4).value)
    ledger_date.append(ws1.cell(row=i, column=5).value)
    ref_text.append(ws1.cell(row=i, column=6).value)


# Convert CBT_keyList int type to string.
# Convert CBT_valueList int/float type to all floats.
CBT_keyListstr = [str(x) for x in CBT_key]
CBT_valueList = [float(i) for i in CBT_value]
CBT_ledgerList = [str(i) for i in ledger_date]
CBT_ledgerListstr = [x[:-6] + '-' + x[-6:-4] + '-' + x[-4:] for x in CBT_ledgerList]

CBT_matchedvalueList = []
CBT_matchedkeyList = []
# Generates the CBT value list that will be matched with the CBT key list.
for i, elem_CBT_list in enumerate(CBT_keyListstr):
    for j, elem_CBT_tuple in enumerate(cbtaccnumlist_tuple):
        if CBT_keyListstr[i] == cbtaccnumlist_tuple[j][0]:
            CBT_matched = cbtaccnumlist_tuple[j][1]
            CBT_matchedkeyList.append(CBT_matched)
            CBT_matchedvalueList.append(CBT_valueList[i])

ref_text_list = []
# Generates the reference text list that will be matched with the CBT value list.
for i, elem_CBT_list in enumerate(CBT_keyListstr):
    for j, elem_CBT_tuple in enumerate(cbtaccnumlist_tuple):
        if CBT_keyListstr[i] == cbtaccnumlist_tuple[j][0]:
            CBT_matched = cbtaccnumlist_tuple[j][1]
            CBT_matchedkeyList.append(CBT_matched)
            ref_text_list.append(ref_text[i])

ledger_date_list = []
# Generates the reference text list that will be matched with the CBT value list.
for i, elem_CBT_list in enumerate(CBT_keyListstr):
    for j, elem_CBT_tuple in enumerate(cbtaccnumlist_tuple):
        if CBT_keyListstr[i] == cbtaccnumlist_tuple[j][0]:
            CBT_matched = cbtaccnumlist_tuple[j][1]
            CBT_matchedkeyList.append(CBT_matched)
            ledger_date_list.append(CBT_ledgerListstr[i])


ref_text_tuple = list(tuple(zip(CBT_matchedvalueList, ref_text_list, ledger_date_list)))
CBT_tuple = list(tuple(zip(CBT_matchedkeyList, ref_text_tuple)))
CBTgrouping = [(k, list(list(zip(*g))[1])) for k, g in groupby(CBT_tuple, itemgetter(0))]

Yardi_value = []
Yardi_key = []
Yardi_date = []

for i in range(7, ws2.max_row):
    try:
        Yardi_key.append(ws2.cell(row=i, column=1).value)
        Yardi_date_str = str(ws2.cell(row=i, column=4).value)
        Yardi_date.append(Yardi_date_str[:10])
        Yardi_value.append(ws2.cell(row=i, column=6).value)
    except ValueError:
        pass

Yardi_keyList = Yardi_key
Yardi_dateList = Yardi_date

Yardi_dateList = [x[5:10] + '-' + x[:4] for x in Yardi_dateList]
# Convert Yardi_valueList int/float type to all floats.
Yardi_valueList = [float(i) for i in Yardi_value]

Yardi_date_trunc = [i[-2:] for i in Yardi_date]
Yardi_date_num = [int(i) for i in Yardi_date_trunc]
Yardi_date_num.sort()


Yardi_date_num_zero = []
for i, elem in enumerate(Yardi_date_num):
    if elem < 10:
        elem = '0' + str(elem)
        Yardi_date_num_zero.append(elem)
    else:
        Yardi_date_num_zero.append(elem)

Yardi_date_num = [str(i) for i in Yardi_date_num_zero]

yardi_days = [Yardi_date_num[i] for i in (0, -1)]
yardi_year = str(Yardi_date[0][:4])
yardi_month = str(Yardi_date[0][-5:7])

wb = Workbook()
# Creating the filename date that includes the Yardi starting and end dates.
Sheet_name = wb.sheetnames

# Creates new Excel file.
root = fr'C:\Users\vpham\PycharmProjects\pythonProject\Programs\CBT Yardi Recon\CBT Yardi Recon'

wb.save(fr'{root} {yardi_month}-{yardi_days[0]}-{yardi_year} to {yardi_month}-{yardi_days[1]}-{yardi_year}.xlsx')
CBT_Yardi_Recon_path = fr'{root} {yardi_month}-{yardi_days[0]}-{yardi_year} to {yardi_month}-{yardi_days[1]}-{yardi_year}.xlsx'
wb3 = xl.load_workbook(CBT_Yardi_Recon_path)
wb4 = xl.load_workbook(CBT_Yardi_Reference_path)
ws3 = wb3.worksheets[0]

# Truncates key at '-' for yardi code key.
Yardi_keyListTrunc = []
for key in Yardi_keyList:
    trunc_key = key.split(' -')[-2]
    Yardi_keyListTrunc.append(trunc_key)

Yardi_matchedkeyList = []
Yardi_matchedvalueList = []
for v, elem_Yardi_list in enumerate(Yardi_keyListTrunc):
    for n, elem_Yardi_tuple in enumerate(yardicodelist_tuple):
        if Yardi_keyListTrunc[v] == yardicodelist_tuple[n][0]:
            Yardi_matchedkey = yardicodelist_tuple[n][1]
            Yardi_matchedkeyList.append(Yardi_matchedkey)
            Yardi_matchedvalueList.append(Yardi_valueList[v])

Yardi_matchedateList = []
for v, elem_Yardi_list in enumerate(Yardi_keyListTrunc):
    for n, elem_Yardi_tuple in enumerate(yardicodelist_tuple):
        if Yardi_keyListTrunc[v] == yardicodelist_tuple[n][0]:
            Yardi_matchedkey = yardicodelist_tuple[n][1]
            Yardi_matchedkeyList.append(Yardi_matchedkey)
            Yardi_matchedateList.append(Yardi_dateList[v])

Yardi_date_Tuple = list(tuple(zip(Yardi_matchedkeyList, Yardi_matchedateList)))
Yardi_Tuple = list(tuple(zip(Yardi_matchedkeyList, Yardi_matchedvalueList)))

Yardigrouping = [(k, list(list(zip(*g))[1])) for k, g in groupby(Yardi_Tuple, itemgetter(0))]
YardiDategrouping = [(k, list(list(zip(*g))[1])) for k, g in groupby(Yardi_date_Tuple, itemgetter(0))]

groupinglist = [CBTgrouping, Yardigrouping, YardiDategrouping]

# Sorts list of tuples, "groups", via bubble sort algorithm.
for groups in groupinglist:
    new_ele = 0
    new_lis_len = len(groups)
    for k in range(0, new_lis_len):
        for i in range(0, new_lis_len - k - 1):
            if groups[i][new_ele] > groups[i + 1][new_ele]:
                new_tem = groups[i]
                groups[i] = groups[i + 1]
                groups[i + 1] = new_tem

wb3 = xl.load_workbook(CBT_Yardi_Recon_path)
sheetpropnames = propertynamelist
sheetpropnames.sort()

for elem in sheetpropnames:
    wb3.create_sheet(elem)

for i, elem in enumerate(CBTgrouping):
    for j, sheets in enumerate(wb3.sheetnames):
        if CBTgrouping[i][0] == wb3.sheetnames[j]:
            wb3.active = wb3[f'{wb3.sheetnames[j]}']
            ws3 = wb3[f'{wb3.sheetnames[j]}']
            # Writes the Property name to the property's sheet.
            ws3.cell(row=1, column=1).value = wb3.sheetnames[j]
            ws3.cell(row=2, column=3).value = "CBT"
            for x, elems in enumerate(CBTgrouping[i][1]):
                # Writes the CBT deposits to CBT column in the property's sheet.
                ws3.cell(row=x + 3, column=3).value = CBTgrouping[i][1][x][0]
                ws3[f'C{x + 3}'].number_format = '$#,##0.00_);($#,##0.00)'
                try:
                    # Writes the associated reference text to Description column in the property's sheet.
                    ws3.cell(row=x + 3, column=8).value = CBTgrouping[i][1][x][1]
                except ValueError:
                    pass
                try:
                    # Writes the associated reference text to Description column in the property's sheet.
                    ws3.cell(row=x + 3, column=2).value = CBTgrouping[i][1][x][2]
                except ValueError:
                    pass

for i, elem in enumerate(YardiDategrouping):
    for j, sheets in enumerate(wb3.sheetnames):
        if YardiDategrouping[i][0] == wb3.sheetnames[j]:
            wb3.active = wb3[f'{wb3.sheetnames[j]}']
            ws3 = wb3[f'{wb3.sheetnames[j]}']
            for x, elems in enumerate(Yardigrouping[i][1]):
                # Writes the Yardi posted date to property's sheet.
                ws3.cell(row=x + 3, column=1).value = YardiDategrouping[i][1][x]

for i, elem in enumerate(cbtaccnumlist_tuple):
    for j, sheets in enumerate(wb3.sheetnames):
        if cbtaccnumlist_tuple[i][1] == wb3.sheetnames[j]:
            wb3.active = wb3[f'{wb3.sheetnames[j]}']
            ws3 = wb3[f'{wb3.sheetnames[j]}']
            # Writes the CBT account number to property's sheet.
            ws3.cell(row=1, column=2).value = int(cbtaccnumlist_tuple[i][0])


for i, elem in enumerate(Yardigrouping):
    for j, sheets in enumerate(wb3.sheetnames):
        if Yardigrouping[i][0] == wb3.sheetnames[j]:
            wb3.active = wb3[f'{wb3.sheetnames[j]}']
            ws3 = wb3[f'{wb3.sheetnames[j]}']
            ws3.cell(row=2, column=1).value = "Posted Date (Yardi)"
            ws3.cell(row=2, column=2).value = "Ledger Date (CBT)"
            ws3.cell(row=2, column=4).value = "Yardi"
            ws3.cell(row=2, column=7).value = "Total Difference"
            ws3.cell(row=3, column=7).value = '=(SUM(E3:E50))-(SUM(F3:F50))'
            ws3.cell(row=2, column=8).value = "Description (CBT)"
            ws3[f'G3'].number_format = '$#,##0.00_);($#,##0.00)'
            for x, elems in enumerate(Yardigrouping[i][1]):
                # Writes the Yardi deposits to property sheet.
                ws3.cell(row=x + 3, column=4).value = Yardigrouping[i][1][x]
                ws3[f'D{x + 3}'].number_format = '$#,##0.00_);($#,##0.00)'
            set_bold('A2:H2')
            ws3['G3'].fill = yellowFill
            red()

for j, sheets in enumerate(wb3.sheetnames):
    wb3.active = wb3[f'{wb3.sheetnames[j]}']
    ws3 = wb3[f'{wb3.sheetnames[j]}']
    ws3.cell(row=2, column=5).value = "Not in Yardi"
    ws3.cell(row=2, column=6).value = "Not in CBT"

    CBT_values = []
    Yardi_values = []

    i = -1
    while True:
        i += 1
        # Reads the CBT deposits back into a list from the property sheet.
        CBT_values.append(ws3.cell(row=i + 3, column=3).value)
        if None in CBT_values:
            break

    i = -1
    while True:
        i += 1
        # Reads the Yardi deposits back into a list from the property sheet.
        Yardi_values.append(ws3.cell(row=i + 3, column=4).value)
        if None in Yardi_values:
            break

    CBT_values = [0 if elem is None else elem for elem in CBT_values]
    Yardi_values = [0 if elem is None else elem for elem in Yardi_values]
    CBT_values.pop()
    Yardi_values.pop()
    # Finds the Set Difference between the two sets.
    CBT_diff = list(set(CBT_values).difference(Yardi_values))
    Yardi_diff = list(set(Yardi_values).difference(CBT_values))

    for x, CBT_diff_cell in enumerate(CBT_diff):
        # Writes the CBT difference to property sheet, "Not in CBT".
        ws3.cell(row=x + 3, column=5).value = CBT_diff_cell
        ws3[f'E{x + 3}'].number_format = '$#,##0.00_);($#,##0.00)'
    for y, Yardi_diff_cell in enumerate(Yardi_diff):
        # Writes the Yardi difference to property sheet, "Not in Yardi".
        ws3.cell(row=y + 3, column=6).value = Yardi_diff_cell
        ws3[f'F{y + 3}'].number_format = '$#,##0.00_);($#,##0.00)'

    set_border(f'A2:H{ws3.max_row - 1}')


del wb3['Sheet']
ws3 = wb3.create_sheet("Summary", 0)
wb3.active = wb3["Summary"]
ws3.cell(row=1, column=1).value = f"Bank Deposits Reconciliation as of {yardi_month}-{yardi_days[1]}-{yardi_year}"
ws3.cell(row=2, column=2).value = "Difference Yardi Vs CBT"
ws3.cell(row=2, column=4).value = "If shows negative deposit(s) is missed in CBT"
ws3.cell(row=3, column=4).value = "If it shows positive deposit(s) are missed in Yardi"
ws3.cell(row=5, column=4).value = "Sum of Difference Yardi Vs CBT"
ws3.cell(row=6, column=4).value = '=(SUM(B3:B1000))'
ws3[f'D6'].number_format = '$#,##0.00_);($#,##0.00)'

for i, propnames in enumerate(sheetpropnames):
    ws3.cell(row=i + 3, column=1).value = propnames
    # Writes the linked differences to summary sheet.
    ws3.cell(row=i + 3, column=2).value = f"='{propnames}'!G3"

for j, sheets in enumerate(wb3.sheetnames):
    wb3.active = wb3[f'{wb3.sheetnames[j]}']
    ws3 = wb3[f'{wb3.sheetnames[j]}']
    for column_cells in ws3.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        # Manually resizes the width of Excel cells.
        if new_column_letter == 'A':
            ws3.column_dimensions['A'].width = 18
        elif new_column_letter == 'B':
            ws3.column_dimensions['B'].width = 18
        elif new_column_letter == 'C':
            ws3.column_dimensions['C'].width = 12
        elif new_column_letter == 'D':
            ws3.column_dimensions['D'].width = 12
        elif new_column_letter == 'G':
            ws3.column_dimensions['G'].width = 16
        else:
            if new_column_length > 0:
                # Auto resizes the width of each other column in every sheet.
                ws3.column_dimensions[new_column_letter].width = new_column_length * 1.23
    set_alignment(f'A3:G{ws3.max_row}', 'right')
    set_alignment(f'B1:B1', 'left')

wb3.active = wb3[f'Summary']
ws3 = wb3[f'Summary']
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 23
ws3.column_dimensions['D'].width = 45
set_border(f'A3:B{ws3.max_row}')
set_bold('A1:B2')
set_bold('D2:D5')
ws3['D6'].fill = yellowFill
red()
for i in range(3, ws3.max_row + 1):
    ws3[f'B{i}'].number_format = '$#,##0.00_);($#,##0.00)'
set_alignment(f'D2:D5', 'left')
wb3.save(CBT_Yardi_Recon_path)

print("------- Program finished running in %s seconds. -------" % round((time.time() - start_time), 2))
