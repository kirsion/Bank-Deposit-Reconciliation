# Place the CBT and Yardi Deposit Excel files into the "Deposit_excel_files" folder, create it if it doesn't exist. The filename has to be "CBT Deposits" and "Yardi Deposits".

import time
import openpyxl as xl
import pandas as pd
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string

start_time = time.time()

CBT_Deposits_path = r'C:\Users\vpham\PycharmProjects\pythonProject\Programs\CBT Yardi Recon\Deposit excel files\CBT Deposits.xlsx'
Yardi_Deposits_path = r'C:\Users\vpham\PycharmProjects\pythonProject\Programs\CBT Yardi Recon\Deposit excel files\Yardi Deposits.xlsx'

wb1 = xl.load_workbook(CBT_Deposits_path)
# Pandas module to sort a column's rows alphabetically or numerically (the column's row first character value has to begin with letters OR numbers, rows cannot contain both).
df = pd.DataFrame(pd.read_excel(CBT_Deposits_path, sheet_name=0))
# Insert the column's title to be sorted in the single quotes.
df.sort_values('Reference Text').to_excel(CBT_Deposits_path)
del df


wb1 = xl.load_workbook(CBT_Deposits_path)
ws1 = wb1.worksheets[0]
ws1.delete_cols(1, 1)
wb1.save(CBT_Deposits_path)
wb1.close()

ReferenceTextList = []
for i in range(0, ws1.max_row):
    ReferenceTextList.append(ws1.cell(row=i+2, column=6).value)
ReferenceTextList = list(filter(None, ReferenceTextList))
ReferenceTextRows = []
for i, ReferenceText in enumerate(ReferenceTextList):
    # Deletes rows that contain these words or strings. Add more conditions to delete with a '' in between and a comma.
    substrings = ['RD   TREAS 310  9101036151', 'CORPORATE XFER FROM', 'WIRE/IN-', 'LPL', 'HYDER PAYROLL']
    if any(x in ReferenceText for x in substrings):
        ReferenceTextRows.append(i+2)
for i in reversed(ReferenceTextRows):
    ws1.delete_rows(i, 1)

wb1.save(CBT_Deposits_path)
wb1.close()

# Pandas module to sort a column's rows alphabetically or numerically (the column's row first character value has to begin with letters OR numbers, rows cannot contain both).
df = pd.DataFrame(pd.read_excel(CBT_Deposits_path, sheet_name=0))
# Insert the column's title to be sorted in the single quotes.
df.sort_values('Account Name').to_excel(CBT_Deposits_path)
del df
wb1 = xl.load_workbook(CBT_Deposits_path)
ws1 = wb1.worksheets[0]
ws1.delete_cols(1, 1)

MIN_WIDTH = 10
for i, column_cells in enumerate(ws1.columns, start=1):
    width = (
        length
        if (length := max(len(str(cell.value)) for cell in filter(None, column_cells)
                          for cell in column_cells)) >= MIN_WIDTH
        else MIN_WIDTH
    )
    ws1.column_dimensions[get_column_letter(i)].width = width

wb1.save(CBT_Deposits_path)
wb1.close()

wb2 = xl.load_workbook(Yardi_Deposits_path)
ws2 = wb2.worksheets[0]
amountList = []
for i in range(0, ws2.max_row-6):
    amountList.append(ws2.cell(row=i+7, column=6).value)
amountRows = []
for i, amount in enumerate(amountList):
    # Deletes negative deposits in Yardi Excel.
    if amount < 0:
        amountRows.append(i+7)
for i in reversed(amountRows):
    ws2.delete_rows(i, 1)

wb2.save(Yardi_Deposits_path)

print("------- Program finished running in %s seconds. -------" % round((time.time() - start_time), 2))
